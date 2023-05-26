from django.shortcuts import get_object_or_404, render, redirect
from openpyxl.utils import get_column_letter 
from firstapp.models import Post,ListGroups,ListStudents,UserProfileInfo
from django.contrib import messages
from firstapp.forms import  UserEditForm, UserFilterForm,UserForm,AddGroupForm,AddStudentsForm,FilterStudentsForm, UserProfileInfoForm,UserProfileInfoEditForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
def home(request):
     
    return render(request, "home.html")
 
def register(request):
    if request.method == "POST":
        form = UserForm(request.POST)
        profile_form = UserProfileInfoForm(request.POST)
        if form.is_valid() and profile_form.is_valid():
            user = form.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            
            profile.save()
            return redirect("login")  # Перенаправление на страницу входа
    else:
        form = UserForm()
        profile_form = UserProfileInfoForm()
    
    context = {
        "form": form,
        "profile_form": profile_form
    }
    return render(request, "register.html", context)



    

@login_required
def user_groups(request):
  
    user = request.user
    
    groups = ListGroups.objects.filter(user=user)
  
    context = {
        "groups": groups
    }
     
    return render(request, "groups.html", context)



def add_group(request):
    if request.method == "POST":
         

        form = AddGroupForm(request.POST)
        if form.is_valid():
            form = form.save(commit=False)
            form.user = request.user
            form.save()
             
            return redirect ("user_groups")
        

    else:
        form = AddGroupForm()
        context = {
            "form":form
        }
        return render ( request, "add_group.html", context)
    


@login_required
def user_students_list(request):
    user = request.user
    students = ListStudents.objects.filter(user=user)

    context = {
        "students": students
    }  
    return render(request, "students_list.html", context)
 
def add_students(request):
    if request.method == "POST":
        form = AddStudentsForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)
            student.user = request.user
            student.save()

            # Отфильтровать доступные группы только для текущего пользователя
            groups = ListGroups.objects.filter(user=request.user)
            student.groups.set(groups)

            form.save_m2m()  # Сохраняем связи "многие ко многим" после сохранения ученика
            return redirect("user_students_list")
    else:
        form = AddStudentsForm()

    # Ограничить выборку групп в форме только доступными группами пользователя
    form.fields['groups'].queryset = ListGroups.objects.filter(user=request.user)

    context = {
        "form": form
    }
    return render(request, "add_students.html", context)


def group_detail(request, group_id):
    group = get_object_or_404(ListGroups, pk=group_id)
    
    # Фильтровать студентов только для текущего пользователя и привязанных к данной группе
    students = group.liststudents_set.filter(user=request.user)

    context = {
        "group": group,
        "students": students
    }
    return render(request, "group_detail.html", context)


# def group_detail(request, group_id):
#     group = get_object_or_404(ListGroups, pk=group_id)
#     students = group.liststudents_set.all()  # Получить студентов, связанных с группой

#     context = {
#         "group": group,
#         "students": students
#     }
#     return render(request, "group_detail.html", context)


def edit_student(request, student_id):
    student = get_object_or_404(ListStudents, pk=student_id)

    # Отфильтровать доступные группы только для текущего пользователя
    groups = ListGroups.objects.filter(user=request.user)

    if request.method == "POST":
        form = AddStudentsForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect('user_students_list')
    else:
        form = AddStudentsForm(instance=student)

    # Ограничить выборку групп в форме только доступными группами
    form.fields['groups'].queryset = groups

    context = {
        'form': form,
        'student': student,
    }
    return render(request, 'edit_student.html', context)




import openpyxl
from django.http import HttpResponse

def export_students_to_excel(request):
    
    # Получение данных всех студентов
    # students = ListStudents.objects.all()
    students = ListStudents.objects.filter(user=request.user)

    # Создание нового файла Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Запись заголовков столбцов
    worksheet.cell(row=1, column=1, value='ФИО')
    worksheet.cell(row=1, column=2, value='Возраст')
    worksheet.cell(row=1, column=3, value='Дата рождения')
    worksheet.cell(row=1, column=4, value='Место обучения')
    worksheet.cell(row=1, column=5, value='ПФДО')
    worksheet.cell(row=1, column=6, value='ФИО родителей')
    worksheet.cell(row=1, column=7, value='Контактные данные')
    worksheet.cell(row=1, column=8, value='Достижения')
    worksheet.cell(row=1, column=9, value='Группа')
    
    # Добавьте остальные заголовки столбцов в соответствии с вашей моделью
    # Установка ширины колонок
    column_widths = [30, 15, 30, 30, 30, 30, 30, 30, 30]  # Ширина каждой колонки
    for i, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = width
    # Запись данных студентов
    row_num = 2
    for student in students:
        worksheet.cell(row=row_num, column=1, value=student.fio)
        worksheet.cell(row=row_num, column=2, value=student.age)
        worksheet.cell(row=row_num, column=3, value=student.birsday)
        worksheet.cell(row=row_num, column=4, value=student.place_learning)
        worksheet.cell(row=row_num, column=5, value=student.sertificat_pfdo)
        worksheet.cell(row=row_num, column=6, value=student.fio_parents)
        worksheet.cell(row=row_num, column=7, value=student.contact_data)
        worksheet.cell(row=row_num, column=8, value=student.atchievments)
        groups = ', '.join([group.name for group in student.groups.all()])
        worksheet.cell(row=row_num, column=9, value=groups)
       
        # Запишите остальные данные студента в соответствии с вашей моделью
        row_num += 1

    # Создание HTTP-ответа с файлом Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    workbook.save(response)

    return response


def delete_student(request, student_id):
    student = get_object_or_404(ListStudents, id=student_id)

    if request.method == "POST":
        student.delete()
        messages.success(request, "Студент успешно удален.")
        return redirect("user_students_list")

    context = {
        "student": student
    }

    return render(request, "delete_student.html", context)



# def filter_students(request):
#     age = request.GET.get("age")
#     gender = request.GET.get("gender")

#     # students = ListStudents.objects.all()
#     students = ListStudents.objects.filter(user=request.user)  # Фильтр студентов текущего пользователя

#     if age:
#         age_range = age.split('-')
#         if len(age_range) == 2:
#             min_age, max_age = int(age_range[0]), int(age_range[1])
#             students = students.filter(age__gte=min_age, age__lte=max_age)

#     if gender:
#         students = students.filter(pol=gender)

#     context = {
#         'students': students,
#         'filter_form': FilterStudentsForm(),
#     }

#     return render(request, 'students_list.html', context)

def filter_students(request):
    age = request.GET.get("age")
    gender = request.GET.get("gender")

    students = ListStudents.objects.filter(user=request.user)  # Фильтр студентов текущего пользователя

    if age:
        age_range = age.split('-')
        if len(age_range) == 2:
            min_age, max_age = int(age_range[0]), int(age_range[1])
            students = students.filter(age__gte=min_age, age__lte=max_age)

    if gender:
        students = students.filter(pol=gender)

    export = request.GET.get("export")  # Проверка наличия параметра "export" в URL

    if export:
        # Создание нового файла Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Запись заголовков столбцов
        worksheet.cell(row=1, column=1, value='ФИО')
        worksheet.cell(row=1, column=2, value='Возраст')
        worksheet.cell(row=1, column=3, value='Дата рождения')
        worksheet.cell(row=1, column=4, value='Место обучения')
        worksheet.cell(row=1, column=5, value='ПФДО')
        worksheet.cell(row=1, column=6, value='ФИО родителей')
        worksheet.cell(row=1, column=7, value='Контактные данные')
        worksheet.cell(row=1, column=8, value='Достижения')
        worksheet.cell(row=1, column=9, value='Группа')


        column_widths = [30, 15, 30, 30, 30, 30, 30, 30, 30]  # Ширина каждой колонки
        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = width

        # Запись данных студентов
        row_num = 2
        for student in students:
            
            worksheet.cell(row=row_num, column=1, value=student.fio)
            worksheet.cell(row=row_num, column=2, value=student.age)
            worksheet.cell(row=row_num, column=3, value=student.birsday)
            worksheet.cell(row=row_num, column=4, value=student.place_learning)
            worksheet.cell(row=row_num, column=5, value=student.sertificat_pfdo)
            worksheet.cell(row=row_num, column=6, value=student.fio_parents)
            worksheet.cell(row=row_num, column=7, value=student.contact_data)
            worksheet.cell(row=row_num, column=8, value=student.atchievments)
            groups = ', '.join([group.name for group in student.groups.all()])
            worksheet.cell(row=row_num, column=9, value=groups)

            row_num += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=filtered_students.xlsx'
        workbook.save(response)

        return response

    context = {
        'students': students,
        'filter_form': FilterStudentsForm(),
        'export_url': request.get_full_path() + '&export=1'  # URL для экспорта данных в Excel
    }

    return render(request, 'students_list.html', context)









def edit_group(request, group_id):
    group = get_object_or_404(ListGroups, pk=group_id)

    if request.method == "POST":
        form = AddGroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            return redirect('group_detail', group_id=group_id)
    else:
        form = AddGroupForm(instance=group)

    context = {
        'form': form,
        'group': group,
    }
    return render(request, 'edit_group.html', context)



def delete_group(request, group_id):
    group = get_object_or_404(ListGroups, pk=group_id)
     

    if request.method == "POST":
        # При удалении группы можно решить, что делать со студентами
        # В данном примере предполагается удаление студентов вместе с группой
        group.liststudents_set.all().delete()

        group.delete()
        return redirect('home')

    context = {
        'group': group,
       
    }
    return render(request, 'delete_group.html', context)


def edit_profile(request):
    user = request.user
    profile = user.profile

    if request.method == 'POST':
        user_form = UserEditForm(request.POST, instance=user)
        profile_form = UserProfileInfoEditForm(request.POST, instance=profile)
         

        if user_form.is_valid() and profile_form.is_valid() :
            user_form.save()
            profile_form.save()

             
            profile.save()

            messages.success(request, 'Профиль успешно обновлен.')
            return redirect('edit_profile')
    else:
        user_form = UserEditForm(instance=user)
        profile_form = UserProfileInfoEditForm(instance=profile)
        
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        
    }
    return render(request, 'edit_profile.html', context)




#Админка
  

from openpyxl.utils import get_column_letter
from openpyxl import Workbook
 
 
from django.http import HttpResponse
 

def user_list(request):
    filter_form = UserFilterForm(request.GET)
    teachers = User.objects.filter(is_superuser=False).select_related('profile')  # Инициализация переменной teachers

    if filter_form.is_valid():
        experience_min = filter_form.cleaned_data['experience_min']
        experience_max = filter_form.cleaned_data['experience_max']
        age_min = filter_form.cleaned_data['age_min']
        age_max = filter_form.cleaned_data['age_max']

        if experience_min:
            teachers = teachers.filter(profile__experience__gte=experience_min)
        if experience_max:
            teachers = teachers.filter(profile__experience__lte=experience_max)
        if age_min:
            teachers = teachers.filter(profile__age__gte=age_min)
        if age_max:
            teachers = teachers.filter(profile__age__lte=age_max)
    
    export = request.GET.get("export")
    if export:
        workbook = Workbook()
        worksheet = workbook.active

        # Запись заголовков столбцов
        worksheet['A1'] = 'Имя'
        worksheet['B1'] = 'Фамилия'
        worksheet['C1'] = 'Направление'
        worksheet['D1'] = 'Стаж(в месяцах)'
        worksheet['E1'] = 'Образование'
        worksheet['F1'] = 'Возраст'
        worksheet['G1'] = 'Место обучения'  
        worksheet['H1'] = 'Достижения'
        worksheet['I1'] = 'Группа(ы)'
        

        # Запись данных преподавателей
        row_num = 2
        for teacher in teachers:
            worksheet.cell(row=row_num, column=1, value=teacher.first_name)
            worksheet.cell(row=row_num, column=2, value=teacher.last_name)
            worksheet.cell(row=row_num, column=3, value=teacher.profile.direction)
            worksheet.cell(row=row_num, column=4, value=teacher.profile.experience)
            worksheet.cell(row=row_num, column=5, value=teacher.profile.education)
            worksheet.cell(row=row_num, column=6, value=teacher.profile.age )
            worksheet.cell(row=row_num, column=7, value=teacher.profile.place_education)
            worksheet.cell(row=row_num, column=8, value=teacher.profile.atchievments)
            groups = ', '.join([group.name for group in teacher.profile.user.listgroups_set.all()])

            worksheet.cell(row=row_num, column=9, value=groups)
             


            row_num += 1
            
            column_widths = [30, 15, 30, 30, 30, 30, 30, 30,30]  # Ширина каждой колонки
        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=filtered_teachers.xlsx'
        workbook.save(response)
        return response

    context = {
        'teachers': teachers,
        'filter_form': filter_form,
        'export_url': request.get_full_path() + '&export=1'
    }

    return render(request, 'admin/user_list.html', context)




import openpyxl
from django.http import HttpResponse

def export_teachers_to_excel(request):
    teachers = User.objects.filter(is_superuser=False, profile__isnull=False)

    # Создание нового файла Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Заголовки столбцов
    sheet['A1'] = 'Имя'
    sheet['B1'] = 'Фамилия'
    sheet['C1'] = 'Направление'
    sheet['D1'] = 'Стаж'
    sheet['E1'] = 'Образование'
    sheet['F1'] = 'Возраст'
    sheet['G1'] = 'Место обучения'  
    sheet['H1'] = 'Достижения'

    # Запись информации о каждом учителе
    for i, teacher in enumerate(teachers, start=2):
        sheet[f'A{i}'] = teacher.first_name
        sheet[f'B{i}'] = teacher.last_name
        sheet[f'C{i}'] = teacher.profile.experience
        sheet[f'D{i}'] = teacher.profile.age

    # Сохранение файла Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=teachers.xlsx'
    workbook.save(response)

    return response



def teacher_group(request, teacher_id, group_id):
    teacher = get_object_or_404(User, id=teacher_id, is_superuser=False, profile__isnull=False)
    group = get_object_or_404(ListGroups, id=group_id, user=teacher)

    # Получить список учеников, принадлежащих к данной группе
    students = ListStudents.objects.filter(groups=group)

    context = {
        'teacher': teacher,
        'group': group,
        'students': students
    }

    return render(request, 'admin/teacher_group.html', context)











 








 

 




 